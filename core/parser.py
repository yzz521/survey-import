# -*- coding: utf-8 -*-
"""解析：门店清单 xlsx/xls/csv + 单店问卷 xlsx/xls/csv。"""
from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl


class UnsupportedFormat(ValueError):
    """不被支持的表格格式。"""


def norm_text(s) -> str:
    """文本规范化：用于题目/选项的跨平台比对。"""
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\xa0", " ").replace("\u3000", " ")
    s = re.sub(r"\s+", "", s)
    s = (s.replace("（", "(").replace("）", ")")
          .replace("，", ",").replace("：", ":").replace("；", ";")
          .replace("“", '"').replace("”", '"')
          .replace("’", "'").replace("‘", "'")
          .replace("－", "-").replace("—", "-"))
    return s.strip()


def read_table(path, *, min_row: int = 1, sheet_index: int = 0,
               ext_hint: str | None = None):
    """把表格（xlsx/xls/csv）读成按行迭代的生成器。

    返回：generator[tuple[Any, ...]]  每行与原表格一致，长度可能不一。
          末尾空行（所有列均为 None/''）会被自动跳过。

    异常：
      UnsupportedFormat   -- 扩展名不在白名单内
      FileNotFoundError   -- 路径不存在
      其他原库错误向上抛
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"找不到文件：{path}")
    ext = (ext_hint or p.suffix).lower().lstrip(".")
    if ext in ("xlsx", "xlsm"):
        return _iter_xlsx(p, min_row=min_row, sheet_index=sheet_index)
    if ext == "csv":
        return _iter_csv(p, min_row=min_row)
    if ext == "xls":
        return _iter_xls(p, min_row=min_row, sheet_index=sheet_index)
    raise UnsupportedFormat(
        f"不支持的文件格式：.{ext}（仅支持 xlsx / xls / csv）")


def _iter_xlsx(path: Path, *, min_row: int, sheet_index: int):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[sheet_index]]
        for row in ws.iter_rows(min_row=min_row, values_only=True):
            if row is None:
                continue
            if all(c is None or (isinstance(c, str) and not c.strip())
                   for c in row):
                continue
            yield tuple(row)
    finally:
        wb.close()


def _iter_csv(path: Path, *, min_row: int = 1):
    """读 CSV：BOM 自动剥离；空行/全空字段行自动跳过。

    `min_row` 为 1-based 行号；之前的行会被跳过。CSV 一般只有 1 行表头，
    所以调用方为 .csv 传 `min_row=2`。

    编码依次尝试 utf-8-sig → gbk → utf-8。
    """
    last_err = None
    for enc in ("utf-8-sig", "gbk", "utf-8"):
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                reader = csv.reader(f)
                for ri, row in enumerate(reader, start=1):
                    if ri < min_row:
                        continue
                    if not row:
                        continue
                    row = tuple(c if c is not None else "" for c in row)
                    if all(str(c).strip() == "" for c in row):
                        continue
                    yield row
            return
        except UnicodeDecodeError as e:
            last_err = e
            continue
    raise UnicodeDecodeError(
        "csv", b"", 0, 1,
        f"无法识别 CSV 编码：{last_err or '尝试 utf-8 / gbk 均失败'}")


def _iter_xls(path: Path, *, min_row: int, sheet_index: int):
    """读旧版 .xls（BIFF8/复合文档）。

    依赖（任选其一即可，优先用已装好的）：
      - xlrd<2.0  + xlrd>=2.0（仅支持 xls，2.0 起不再支持；当前 2.x 安装包对 .xls 会抛错）
      - pandas >= 1.0（自带 xlrd 引擎，可读 .xls）
    都没装的话给清晰的安装指引。
    """
    try:
        import xlrd  # noqa: F401
    except Exception:
        try:
            import pandas as pd  # noqa: F401
        except Exception:
            raise UnsupportedFormat(
                "读取 .xls 需要 xlrd 或 pandas。请执行：\n"
                "  pip install 'xlrd==1.3.2'\n"
                "或\n"
                "  pip install pandas"
            )
    try:
        import xlrd
        wb = xlrd.open_workbook(str(path), formatting_info=False)
        sh = wb.sheet_by_index(sheet_index)
        for ri in range(max(0, min_row - 1), sh.nrows):
            row = sh.row_values(ri)
            if all(c is None or (isinstance(c, str) and not c.strip())
                   for c in row):
                continue
            yield tuple(row)
        return
    except Exception as xlrd_err:
        xlrd_msg = str(xlrd_err)
    # pandas 兜底（pandas 默认会用 xlrd 引擎，2.0+ 同样读不了 .xls；它失败时给出明确提示）
    try:
        import pandas as pd
        df = pd.read_excel(str(path), sheet_name=sheet_index,
                           header=None, engine="xlrd", dtype=str)
        for _, row in df.iterrows():
            row = tuple("" if pd.isna(v) else v for v in row.tolist())
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            yield row
        return
    except Exception as pd_err:
        raise UnsupportedFormat(
            "无法读取 .xls 文件（可能由旧版 Office 生成或二进制损坏）。\n"
            f"xlrd 错误：{xlrd_msg}\n"
            f"pandas 错误：{pd_err}\n"
            f"建议：在 Excel/Sheets 里「另存为 .xlsx 或 .csv」后再导入。"
        )


def list_supported_files(directory) -> list[Path]:
    """扫描一个目录，返回所有支持的表格文件路径（按文件名排序）。

    同时识别 .xlsx, .xlsm, .xls, .csv。文件名以 ~$ 开头的临时锁文件会被忽略。
    """
    out: list[Path] = []
    for ext in ("xlsx", "xlsm", "xls", "csv"):
        for p in sorted(Path(directory).glob(f"*.{ext}")):
            if p.name.startswith("~$"):
                continue
            out.append(p)
    return out


# ---------------------------------------------------------------- 门店清单

@dataclass
class StoreTask:
    seq: int
    record_type: str
    store_code: str
    store_name: str
    brand: str
    region: str
    province: str
    city: str
    address: str
    remark: str
    project: str
    wave: str
    url: str
    access_code: str
    link_status: str
    tenant_key: str = ""
    public_id: str = ""
    questionnaire_path: str = ""


# 表头所在行（1-based）—— 仅作 fallback 兜底
LIST_HEADER_ROW = 4
COLS = {
    "seq": 0, "record_type": 1, "store_code": 2, "store_name": 3,
    "brand": 4, "region": 5, "area_mgr": 6, "ops_mgr": 7,
    "province": 8, "city": 9, "address": 10, "remark": 11,
    "project": 12, "wave": 13, "exec_item": 14,
    "url": 15, "access_code": 16, "link_status": 17, "src_row": 18,
}


def _detect_header_row(rows, keywords: tuple[str, ...], *,
                       lookahead: int = 12) -> tuple[int, int]:
    """在前 `lookahead` 行里找到「同时含所有 keywords」的那行作为表头。

    返回 (header_row_1based, data_start_1based)；找不到则 (0, 0)。
    """
    if not rows:
        return 0, 0
    if not hasattr(rows, "__iter__") or isinstance(rows, (list, tuple)):
        seq = list(rows)
    else:
        seq = list(rows)
    for ri, row in enumerate(seq[:lookahead], start=1):
        joined = "  ".join("" if c is None else str(c) for c in row)
        if all(kw in joined for kw in keywords):
            return ri, ri + 1
    return 0, 0


def parse_link(url: str) -> tuple[str, str]:
    """从外链提取 (tenant_key, public_id)。

    支持 /external/<tenant>/<publicId> 与 /report/<tenant>/<publicId>。
    """
    if not url:
        return "", ""
    parts = [p for p in str(url).split("?")[0].split("#")[0].split("/") if p]
    if len(parts) >= 2:
        return parts[-2], parts[-1]
    return "", ""


def parse_store_list(path: str | Path) -> list[StoreTask]:
    """解析门店清单（xlsx / xls / csv）。

    xlsx/xls 默认 header 在第 4 行，CSV 默认 header 在第 1 行；
    但很多人从 Excel「另存为 CSV」会保留 metadata 行，所以先做自动检测：
    找到第一个同时含「门店编号 / 记录类型 / 访问码 / URL」其中两个的行，作为表头。
    找不到时回退到对应扩展名的固定 start_row（兜底）。
    """
    ext = Path(path).suffix.lower()
    rows = list(read_table(path, min_row=1))
    if not rows:
        return []
    _h, data_start = _detect_header_row(
        rows,
        keywords=("门店编号", "记录类型"),
    )
    if not data_start:
        data_start = 2 if ext == ".csv" else LIST_HEADER_ROW + 1
    tasks: list[StoreTask] = []
    for ri, row in enumerate(rows, start=1):
        if ri < data_start:
            continue
        # 长度兜底：CSV/xls 取出来可能比 19 列短，按索引取 None
        def cell(idx):
            return row[idx] if idx < len(row) else None
        code = cell(COLS["store_code"])
        url = cell(COLS["url"])
        if not code or not url:
            continue
        tenant, pid = parse_link(url)
        tasks.append(StoreTask(
            seq=int(cell(COLS["seq"])) if isinstance(cell(COLS["seq"]), (int, float)) else len(tasks) + 1,
            record_type=str(cell(COLS["record_type"]) or ""),
            store_code=str(code).strip(),
            store_name=str(cell(COLS["store_name"]) or "").strip(),
            brand=str(cell(COLS["brand"]) or ""),
            region=str(cell(COLS["region"]) or ""),
            province=str(cell(COLS["province"]) or ""),
            city=str(cell(COLS["city"]) or ""),
            address=str(cell(COLS["address"]) or ""),
            remark=str(cell(COLS["remark"]) or ""),
            project=str(cell(COLS["project"]) or ""),
            wave=str(cell(COLS["wave"]) or ""),
            url=str(url).strip(),
            access_code=str(cell(COLS["access_code"]) or "").strip(),
            link_status=str(cell(COLS["link_status"]) or ""),
            tenant_key=tenant,
            public_id=pid,
        ))
    return tasks


# ---------------------------------------------------------------- 问卷目录

def index_questionnaire_dir(d: str | Path) -> dict[str, Path]:
    """扫描问卷目录，返回 {门店编号: 文件路径}。

    支持扩展名：xlsx / xls / csv。
    文件名形如 Y10018208-暴龙专卖宁德万达店.xlsx。
    """
    out: dict[str, Path] = {}
    for p in list_supported_files(d):
        m = re.match(r"^([A-Za-z0-9]+)[\-—_]+.*$", p.stem)
        if m:
            out.setdefault(m.group(1).upper(), p)
    return out


# ---------------------------------------------------------------- 单店问卷

@dataclass
class ParsedItem:
    """问卷中的一行实体。"""
    raw_question: str
    answer: str | None = None
    options: list[str] = field(default_factory=list)
    excel_type: str | None = None
    row: int = 0

    @property
    def is_question(self) -> bool:
        """题目行 = 有答案，或标注了题型。分组标题两者皆无。"""
        return bool(self.answer) or bool(self.excel_type)


SURVEY_DATA_START = 4


def parse_survey_file(path: str | Path) -> list[ParsedItem]:
    """解析单店问卷（xlsx/xls/csv）。

    列位置：B 题目 / C 答案 / D 题型。
    数据起始行：xlsx/xls 第 4 行（此项目固定）；CSV 一般第 2 行。
    但从 Excel 直接导出的 CSV 经常把前 2 行说明性文字也带过来，所以
    先自动找含「题目描述 + 题目类型」的行作为表头，再从下一行起解析。
    """
    ext = Path(path).suffix.lower()
    rows = list(read_table(path, min_row=1))
    _h, data_start = _detect_header_row(
        rows,
        keywords=("题目描述", "题目类型"),
    )
    if not data_start:
        data_start = 2 if ext == ".csv" else SURVEY_DATA_START
    items: list[ParsedItem] = []
    cur: ParsedItem | None = None
    for ri, row in enumerate(rows, start=1):
        if ri < data_start:
            continue
        b = row[1] if len(row) > 1 else None
        c = row[2] if len(row) > 2 else None
        d = row[3] if len(row) > 3 else None
        if b is not None and str(b).strip():
            if cur:
                items.append(cur)
            cur = ParsedItem(
                raw_question=str(b).strip(),
                answer=str(c).strip() if c is not None and str(c).strip() else None,
                excel_type=str(d).strip() if d is not None and str(d).strip() else None,
                row=ri,
            )
        elif cur is not None and c is not None and str(c).strip():
            cur.options.append(str(c).strip())
    if cur:
        items.append(cur)
    return items
